from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from ler_excel.ler_plan_excel import ler_planilha_excel_ncm_st
from compara_ncm.compara_ncm_planilhas import ncm_produto_x_ncm_alteradas
import os
import csv

INDICE_NCM = 15

# ler a planilha das NCMs ST
lista_ncm = ler_planilha_excel_ncm_st()


# ler csv produtos
lista_produtos_alterados = []
with open('planilhas' + os.sep + 'produtos.csv') as produtos:
    for produto in csv.reader(produtos, delimiter=';'):

        # print(produto[INDICE_NCM])
        if len(produto[INDICE_NCM]) == 7:
            produto[INDICE_NCM] = '0' + produto[INDICE_NCM]

        if len(produto[INDICE_NCM]) == 8:
            eh_igual = ncm_produto_x_ncm_alteradas(
                produto[INDICE_NCM], lista_ncm)
            if eh_igual:
                produto.insert(76, 'True')
            else:
                produto.insert(76, 'False')

        lista_produtos_alterados.append(produto)

# criar nova planilha para colocar resultado
pasta_produtos = Workbook()
del pasta_produtos['Sheet']
pasta_produtos.create_sheet('produtos')
planilha_produtos = pasta_produtos['produtos']

# inserir resultados na planilha
for produtos_alterados in lista_produtos_alterados:
    planilha_produtos.append(produtos_alterados)


# ler planilha e pintar resultados positivos
planilha_produtos = pasta_produtos['produtos']
linha = 1
for produto in planilha_produtos:
    if planilha_produtos[f'BX{linha}'].value == 'True':
        planilha_produtos[f'P{linha}'].fill = PatternFill(
            "solid", start_color="FFFF00")
    linha += 1

pasta_produtos.save('produtos_analisados.xlsx')
